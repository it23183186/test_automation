#!/usr/bin/env python3
"""
Script to populate the Excel file with 50 negative test cases
Registration: IT23183186
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# Test cases data - 50 negative test cases
test_cases = [
    # (TC_ID, Input_Length, Input, Expected_Output)
    ("Neg_0001", "S", "oyata wadida kiyala ahuwa?", "ඔයාට වැඩිද කියලා ඇහුවා?"),
    ("Neg_0002", "M", "mata eyata kawadda katha karanna puluwan ?", "මට එයාට කවදාද කතා කරන්න පුළුවන්?"),
    ("Neg_0003", "S", "Thaththaa eka poddak karanna.", "තාත්තා එක පොඩ්ඩක් කරන්න."),
    ("Neg_0004", "M", "mata oyage number eka denna kiyala kiwwa, ita passe WhatsApp  karannam.", "මට ඔයාගේ number එක දෙන්න කියලා කිව්වා, ඊට පස්සේ WhatsApp කරන්නම්."),
    ("Neg_0005", "S", "ayubowan! suba upadn dinayak wewa!", "ආයුබෝවන්! සුභ උපන් දිනයක් වේවා!"),
    ("Neg_0006", "M", "hello oyala hamoma hodin innawa neda.", "හෙලෝ, ඔයාලා හාමොම හොඳින් ඉන්නවා නේද?"),
    ("Neg_0007", "S", "oyata poddak balanna puluwannam.", "ඔයාට පොඩ්ඩක් බලන්න පුළුවන්නම්."),
    ("Neg_0008", "M", "oyata lesiyata eka mama dora lagata genath dennam ", "ඔයාට ලේසියට එක මම දොර ළඟට ගෙනත් දෙන්නම්."),
    ("Neg_0009", "S", "ow wade iwara una", "ඔව් වැඩේ ඉවර උනා."),
    ("Neg_0010", "M", "hari mama eka hoyala balala karannam", "හරි, මම එක හොයලා බලලා කරන්නම්."),
    ("Neg_0011", "S", "Eka loku udwwak oyata godak godak pin.", "එක ලොකු උදව්වක් ඔයාට ගොඩක් ගොඩක් පිං."),
    ("Neg_0012", "M", "ane, hari hari mama eka karala inne oya karadra wenna one na.", "අනේ, හරි හරි, මම එක කරලා ඉන්නේ, ඔයා කරදර වෙන්න ඕනේ නෑ."),
    ("Neg_0013", "S", "mata eka denna... oyatath balanna puluwan kwadda?", "මට එක දෙන්න... ඔයාත් බලන්න පුළුවන් කවදාද?"),
    ("Neg_0014", "M", "ane! mama ayata katha karannam... mata samawenna!", "අනේ! මම අයාට කතා කරන්නම්... මට සමාවෙන්න!"),
    ("Neg_0015", "S", "mama oyat passe amathumak dennam.", "මම ඔයාත් පස්සේ ආමන්ත්‍රණයක් දෙන්නම්."),
    ("Neg_0016", "M", "api ape pasalata garu karanawa e wageama guruwaruntath.", "අපි අපේ පාසලට ගරු කරනවා, ඒ වගේම ගුරුවරුන්ටත්."),
    ("Neg_0017", "S", "aeyage password eka forget karada ?", "ඇයගෙ password එක forget කරාද?"),
    ("Neg_0018", "M", "oyage document eka email karannam, then signature eka add karanna.", "ඔයාගේ document එක email කරන්නම්, then signature එක add කරන්න."),
    ("Neg_0019", "S", "api as soon as possible yamuda , mata vehesai...", "අපි as soon as possible යමුද? , මට වෙහෙසයි..."),
    ("Neg_0020", "M", "mama office awe, kiyanna feeling tired very bad. oyata  dhairyamath  dennam.", "මම office ආවේ, කියන්න feeling tired very bad. ඔයාට ධෛයර්මත් දෙන්නම්."),
    ("Neg_0021", "S", "oyage account eken logout unada?", "ඔයාගේ account එකෙන් logout උනාද?"),
    ("Neg_0022", "M", "mata download karanna puluwan keeyatada? file format eka CSV da thiynne?", "මට download කරන්න පුළුවන් කීයටද? file format එක CSV ද තියෙන්නේ?"),
    ("Neg_0023", "S", "mata Instagram message ekak ewanna eka ithamath ikmanata.", "මට Instagram message එකක් එවන්න, එක ඉතාමත් ඉක්මනට."),
    ("Neg_0024", "M", "Telegram eken notification awa thamai? Gmail check karada? mata notification ekat reply damme na.", "Telegram එකෙන් notification ආවා තමයි. Gmail check කරාද? මට notification එකට reply දාන්නේ නෑ."),
    ("Neg_0025", "S", "Mama COVID test eka kara.", "මම COVID test එක කරා."),
    ("Neg_0026", "M", "oyage ID eka copy karada? PDF file eka convert karanna puluwanda?", "ඔයාගේ ID එක copy කරාද? PDF file එක convert කරන්න පුළුවන්ද?"),
    ("Neg_0027", "S", "mata university eke exam danna kalin study karanna one .", "මට university එකේ exam දෙන්න කලින් study කරන්න ඕනේ."),
    ("Neg_0028", "M", "ape lab ekedi project ekak gahanna kalin demo ekak karanna wei apita.", "අපේ lab එකේදී project එකක් ගහන්න කලින් demo එකක් කරන්න වෙයි අපිට."),
    ("Neg_0029", "S", "api Galle Trip ekak yanne neddthida?", "අපි Galle Trip එකක් යන්නේ නැද්ද ?"),
    ("Neg_0030", "M", "ada hawasata api Colombo Fort walata shopping karamu? Dematagoda kochchiye yamuda?", "අද හවසට අපි Colombo Fort වලට shopping කරමු. Dematagoda කෝච්චියේ යමුද?"),
    ("Neg_0031", "S", "Sachini meeting ekedi baya weida?", "සාවිත්‍රී meeting එකේදී බය වෙයිද?"),
    ("Neg_0032", "M", "Pradeep nui  Dilshan nui dennama project eka hondata kara. Jayasena eko pasal eka inna.", "ප්‍රදීප් සහ දිල්ශාන් දෙන්නම project එක හොඳට කරා. ජයසේන ඒකෝ පාසල එකේ ඉන්නා."),
    ("Neg_0033", "S", "mama exam ekedi 75n 100yama aragena.", "මම exam එකක්දී 75න් 100යම අරකගන."),
    ("Neg_0034", "M", "api 5 dena watayak yamuda? 2026 awurudhe ape plan wlata anuwa .", "අපි 5 දෙනා වටේ යමුද? 2026 අවුරුද්දේ අපේ plan වලට අනුව."),
    ("Neg_0035", "S", "mata USD 100 denna puluwan katada?", "මට USD 100 දෙන්න පුළුවන් කාටද?"),
    ("Neg_0036", "M", "oyage salary eka Rs. 80000 k kiyanne me kale hatiyata godak watinawa.", "ඔයාගේ salary එක Rs. 80000 කියන්නේ මේ කාලේ හැටියට ගොඩක් වටිනවා."),
    ("Neg_0037", "S", "meeting eka 2:45PM ekata schedule karada?", "meeting එක 2:45PM එකට schedule කරාද?"),
    ("Neg_0038", "M", "heta 08:30 AM ekata enna karunakarala.9.00 AM walata meeting eka thiayanawa.", "හෙට 08:30 AM එකට එන්න කරුණාකරලා. 9:00 AM වලට meeting එක තියෙනවා."),
    ("Neg_0039", "S", "appointment eka 2026-03-15 walata booking karada?", "appointment එක 2026-03-15 වලට booking කරාද?"),
    ("Neg_0040", "M", "next holiday eka 5th May wala thiyanawa. birthday eka 6/5/2026 ekata celebrate karannam.", "next holiday එක 5th May වල තියෙනවා. birthday එක 6/5/2026 එකට celebrate කරන්නම්."),
    ("Neg_0041", "S", "mama 2kg withara bara adu wela .", "මම 2kg විතර බර අඩු වෙලා."),
    ("Neg_0042", "M", "e hariye pare 5.5m witharai hondata thiyenne. temperature 36.5 celsius walata thibba.", "එ හදිසි පාරේ 5.5m විතරයි හොඳට තියෙන්නේ. temperature 36.5 celsius වලට තිබ්බා."),
    ("Neg_0043", "S", "sirawata machan, uba wedin karanava natha?", "සිරාවට චේ, උබ වැඩින් කරනවා නැතා?"),
    ("Neg_0044", "M", "bro eka kochchrada? mama eke set wela inna karunakarala. uba mara wadane karanne!", "bro එක කොච්චරද? මම ඒකේ set වෙලා ඉන්න කරුණාකරලා. උබ මාර වැඩනේ කරන්නේ!"),
    ("Neg_0045", "S", " me link eka balannako eka hodai danumata yamak: https://youtube.com", "මේ link එක බලන්නකො, එක හොඳයි: https://youtube.com"),
    ("Neg_0046", "M", "website ekata username dulandi@user123 kiyala password danna . email: contact@domain.com evanna puluwanda?", "website එකට username dulandi@user123 කියලා password දාන්න. email: contact@domain.com එවන්න පුළුවන්ද?"),
    ("Neg_0047", "S", "mata happy 😊 oyatath sathutuda?", "මම happy 😊 ඔයාත් සතුටුද?"),
    ("Neg_0048", "M", "uba very mahansie patai 😴 poddak rest karanna. 🎯 target eka achieve karannam 💪", "උබ very මහන්සි පාටයි 😴 පොඩ්ඩක් rest කරන්න. 🎯 target එක achieve කරන්නම් 💪"),
    ("Neg_0049", "L", "api semester ekedi 4 subject thiyenne. Science, Maths, IT saha English. mama Science paper ekedi 85 arakagana. oyage subject thiyenna ada exam dates 2026-04-15, 2026-04-18, 2026-04-20, 2026-04-22 ayiya. Zoom eken study group ekak karamu. #ExamPrep #Study", "අපි semester එකේදී 4 subject තියෙන්නේ. Science, Maths, IT සහ English. මම Science paper එකේදී 85 අරගෙනා. ඔයාගේ subject තියෙන්නේ, අද exam dates 2026-04-15, 2026-04-18, 2026-04-20, 2026-04-22 අයියා. Zoom එකෙන් study group එකක් කරමු. #ExamPrep #Study"),
    ("Neg_0050", "L", "machan mage graduation ceremony eka 2026-05-10  thiyanawa ! #Proud #Family 🎓 oyata 7:30AM ekata awe kalin call karannam. mama blue shirt ekat adinne, oyata kawadda? Havelock City reception ekata yamuda? oyage family samaga venue ekat pahala inan ..mama  transport arrange karannum. mata Rs.50000 withara budget thiyenne. 😅", "මචං මගේ graduation ceremony එක 2026-05-10 තියෙනවා! #Proud #Family 🎓 ඔයාට 7:30AM එකට ආවේ කලින් call කරන්නම්. මම blue shirt එකට ඇදිනේ, ඔයාටත් කවදාද? Havelock City reception එකට යමුද? ඔයාගේ family සමඟ venue එකට පහලා ඉන්නා, මම transport arrange කරන්නම්. මට Rs.50000 විතර budget තියෙන්නේ. 😅"),
]

def populate_excel(excel_path, test_cases_data):
    """Populate Excel file with test cases"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Unmerge all cells to avoid read-only issues
    for merged_cell_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell_range))
    
    # Clear existing data (keep headers in row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    
    # Populate test cases starting from row 2
    for idx, (tc_id, input_length, input_text, expected_output) in enumerate(test_cases_data, start=2):
        ws[f'A{idx}'] = tc_id
        ws[f'B{idx}'] = input_length
        ws[f'C{idx}'] = input_text
        ws[f'D{idx}'] = expected_output
        # Leave Actual output and Status empty for automation to fill
        ws[f'E{idx}'] = None
        ws[f'F{idx}'] = None
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 12
    
    # Set text wrapping for better readability
    for row in ws.iter_rows(min_row=2, max_row=len(test_cases_data)+1, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(excel_path)
    print(f"* Successfully populated {len(test_cases_data)} test cases into Excel file")
    print(f"* File saved: {excel_path}")

if __name__ == "__main__":
    excel_file = "IT23183186_Assignment 1 - Test cases.xlsx"
    populate_excel(excel_file, test_cases)