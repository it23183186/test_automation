#!/usr/bin/env python3
"""
Automated Test Suite for Chat Sinhala Transliteration
Registration: IT23183186
"""
import argparse
import openpyxl
import re
import time
import os
from playwright.sync_api import sync_playwright
import sys

def safe_print(s):
    try:
        print(s)
    except UnicodeEncodeError:
        # Fallback for consoles that don't support the characters
        print(s.encode('ascii', 'replace').decode('ascii'))

def run_automation(excel_path, url, wait_ms, type_delay, slow_mo, save_every, headless, keep_open):
    safe_print(f"[*] Starting Automation...")
    safe_print(f"[*] Excel File: {excel_path}")
    safe_print(f"[*] Target URL: {url}")
    
    if not os.path.exists(excel_path):
        print(f"[!] Error: Excel file not found at {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Identify columns
    headers = [cell.value for cell in ws[1]]
    try:
        id_col = headers.index("Test Case ID") + 1
        input_col = headers.index("Input") + 1
        expected_col = headers.index("Expected output") + 1
        actual_col = headers.index("Actual output") + 1
        status_col = headers.index("Status") + 1
    except ValueError as e:
        safe_print(f"[!] Error: Could not find required columns in Excel headers. {e}")
        safe_print(f"Found headers: {headers}")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(60000)
        
        safe_print(f"[*] Navigating to {url}...")
        page.goto(url, wait_until="networkidle")
        
        # Locators
        input_area = page.locator('textarea[placeholder*="English"]').first
        output_area = page.locator('textarea[placeholder*="Sinhala"]').first
        trans_btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
        
        total_rows = ws.max_row
        processed = 0
        
        for row in range(2, total_rows + 1):
            tc_id = ws.cell(row=row, column=id_col).value
            input_text = ws.cell(row=row, column=input_col).value
            expected_text = ws.cell(row=row, column=expected_col).value
            existing_actual = ws.cell(row=row, column=actual_col).value
            
            if not tc_id or not input_text:
                continue
            
            # Skip if already processed
            if existing_actual and ws.cell(row=row, column=status_col).value != "ERROR":
                safe_print(f"[*] [{tc_id}] Skipping (already processed)")
                continue
                
            safe_print(f"[*] [{tc_id}] Processing: {input_text[:40]}...")
            
            try:
                # Clear and Type
                input_area.fill("")
                input_area.type(input_text, delay=type_delay)
                
                # Click Transliterate
                trans_btn.click()
                
                # Wait for response
                time.sleep(wait_ms / 1000)
                
                # Get actual output
                actual_output = output_area.input_value()
                
                # Record result
                ws.cell(row=row, column=actual_col).value = actual_output
                status = "PASS" if actual_output.strip() == expected_text.strip() else "FAIL"
                ws.cell(row=row, column=status_col).value = status
                
                safe_print(f"    -> Status: {status}")
                processed += 1
                
                # Save periodically
                if processed % save_every == 0:
                    wb.save(excel_path)
                    
            except Exception as e:
                safe_print(f"    [!] Error processing row {row}: {e}")
                ws.cell(row=row, column=status_col).value = "ERROR"
                wb.save(excel_path)

        wb.save(excel_path)
        safe_print(f"\n[*] Automation Complete!")
        safe_print(f"[*] Total cases processed: {processed}")
        safe_print(f"[*] Results saved to: {excel_path}")
        
        if not keep_open:
            browser.close()
        else:
            safe_print("[*] Keeping browser open as requested. Press Ctrl+C to terminate.")
            while True:
                time.sleep(10)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Singlish Transliteration Automation")
    parser.add_argument("--excel", default="IT23183186_Assignment 1 - Test cases.xlsx", help="Path to Excel file")
    parser.add_argument("--url", default="https://www.pixelssuite.com/chat-translator", help="Target URL")
    parser.add_argument("--wait-ms", type=int, default=5000, help="Wait time after click (ms)")
    parser.add_argument("--type-delay", type=int, default=80, help="Delay between keystrokes (ms)")
    parser.add_argument("--slow-mo", type=int, default=200, help="Playwright slow-mo (ms)")
    parser.add_argument("--save-every", type=int, default=1, help="Save frequency")
    parser.add_argument("--headless", action="store_true", default=False, help="Run in headless mode")
    parser.add_argument("--keep-open", action="store_true", default=False, help="Keep browser open after tests")
    
    args = parser.parse_args()
    
    run_automation(
        args.excel, args.url, args.wait_ms, args.type_delay, 
        args.slow_mo, args.save_every, args.headless, args.keep_open
    )
