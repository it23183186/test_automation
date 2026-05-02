#!/usr/bin/env python3
"""
Add Singlish input type columns to Excel file
Registration: IT23183186
"""
import openpyxl
from openpyxl.styles import Alignment

# Test case categorization with input types and evidence
test_case_mapping = {
    2: ("Question forms", "Question with 'weddi' spelling variant"),
    3: ("Question forms", "Multi-part question with multiple clauses"),
    4: ("Command forms", "Command with simple imperative"),
    5: ("Command forms, Platform/App Names", "Command with English word insertions (WhatsApp)"),
    6: ("Greetings", "Greeting with spelling variants"),
    7: ("Greetings", "Formal greeting with multiple clauses"),
    8: ("Requests", "Simple request with conditional"),
    9: ("Requests, Isolated English Word Insertions", "Request with English word insertions (please)"),
    10: ("Responses", "Positive response with repetition"),
    11: ("Responses", "Multiple part response"),
    12: ("Repeated Words", "Repeated adjective with warning"),
    13: ("Repeated Words", "Repeated pronoun in phrase"),
    14: ("Inputs with Punctuation Marks", "Ellipsis usage in question"),
    15: ("Inputs with Punctuation Marks", "Multiple punctuation marks (!...)"),
    16: ("Romanization / Spelling Variants", "'mang' instead of 'man'"),
    17: ("Romanization / Spelling Variants", "Multiple spelling variant forms"),
    18: ("Isolated English Word Insertions in Singlish", "Single English technical term (password)"),
    19: ("Isolated English Word Insertions in Singlish", "Multiple isolated English words"),
    20: ("Multi-Word English Phrases in Singlish", "English phrase embedded (as soon as possible)"),
    21: ("Multi-Word English Phrases in Singlish", "Multiple English phrases (feeling tired very bad)"),
    22: ("English Digital Terms in Singlish", "Digital term 'account' and 'logout'"),
    23: ("English Digital Terms in Singlish", "Multiple digital/technical terms (download, CSV)"),
    24: ("Platform/App Names in Singlish", "App name embedded (Instagram)"),
    25: ("Platform/App Names in Singlish", "Multiple app names (Telegram, Gmail)"),
    26: ("English Abbreviations/Acronyms in Singlish", "Health acronym (COVID)"),
    27: ("English Abbreviations/Acronyms in Singlish", "Multiple acronyms (ID, PDF)"),
    28: ("English Clipped Forms in Singlish", "Clipped form 'uni'"),
    29: ("English Clipped Forms in Singlish", "Multiple clipped forms (lab, demo)"),
    30: ("Place Names Embedded in Singlish", "Single place name (Galle)"),
    31: ("Place Names Embedded in Singlish", "Multiple place names (Colombo Fort, Dematagoda)"),
    32: ("Person Names Embedded in Singlish", "Single person name (Sachini)"),
    33: ("Person Names Embedded in Singlish", "Multiple person names (Pradeep, Dilshan, Jayasena)"),
    34: ("Inputs with Numbers and Numeric Suffixes", "Score with numeric suffix (75n 100yama)"),
    35: ("Inputs with Numbers and Numeric Suffixes", "Multiple numeric forms (5 dina, 2025)"),
    36: ("Inputs with Currency", "Foreign currency (USD)"),
    37: ("Inputs with Currency", "Multiple currency types (Rs., GBP)"),
    38: ("Inputs with Time Formats", "Time with period notation (2:45PM)"),
    39: ("Inputs with Time Formats", "Multiple time formats (08:30 AM, 6.00 PM)"),
    40: ("Inputs with Dates", "ISO date format (2026-03-15)"),
    41: ("Inputs with Dates", "Multiple date formats (5th May, 14/02/2026)"),
    42: ("Inputs with Unit of Measurements", "Weight measurement (2kg)"),
    43: ("Inputs with Unit of Measurements", "Multiple measurement types (5.5m, 36.5celsius)"),
    44: ("Inputs with Slang and Casual Phrasing", "Casual greeting with slang (sirawata, machan)"),
    45: ("Inputs with Slang and Casual Phrasing", "Multiple slang expressions (bro, uke set)"),
    46: ("Online Identifiers in Singlish", "URL with path (https://youtube.com/@channel)"),
    47: ("Online Identifiers in Singlish", "Multiple online identifiers (@user123, email)"),
    48: ("Inputs Containing Emojis", "Simple emoji in middle (😊)"),
    49: ("Inputs Containing Emojis", "Multiple emojis (😴, 🎯, 💪)"),
    50: ("Multiple types combined", "Long input with numbers, dates, app names, URLs, hashtags"),
    51: ("Multiple types combined", "Very long input with emojis, dates, currency, time, place names"),
}

def add_input_type_columns(excel_path, mapping):
    """Add Singlish input type columns to Excel"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Add new column headers in columns 7 and 8
    ws.cell(row=1, column=7).value = "Singlish input types covered"
    ws.cell(row=1, column=8).value = "Evidence or rationale for the input type covered"
    
    # Populate the columns based on mapping
    for row_num, (input_types, evidence) in mapping.items():
        ws.cell(row=row_num, column=7).value = input_types
        ws.cell(row=row_num, column=8).value = evidence
    
    # Adjust column widths
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 70
    
    # Set text wrapping and alignment for readability
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(excel_path)
    print(f"* Successfully added Singlish input type columns to Excel file")
    print(f"* File saved: {excel_path}")
    print(f"* Columns added:")
    print(f"  - Column G: Singlish input types covered")
    print(f"  - Column H: Evidence or rationale for the input type covered")

if __name__ == "__main__":
    excel_file = "IT23183186_Assignment 1 - Test cases.xlsx"
    add_input_type_columns(excel_file, test_case_mapping)
