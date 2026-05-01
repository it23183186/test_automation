#!/usr/bin/env python3
"""
Script to populate the Excel file with 50 negative test cases
Registration: IT23183186
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# Test cases data - 50 negative test cases
test_cases = [
    # (TC_ID, Input_Length, Input, Expected_Output)
    ("Neg_0001", "S", "oyata weddi kiyala kiwwa?", "ඔයාට වැඩ්ඩි කියලා කිව්වා?"),
    ("Neg_0002", "M", "mata eyata kawadda katha karannam? wagena?", "මට එයාට කවද්ද තා කරන්නම්? වගෙනා?"),
    ("Neg_0003", "S", "amma eka podak karanna.", "අම්මා එක පඩාක් කරන්න."),
    ("Neg_0004", "M", "mata oyage number eka dapana kiyala, then WhatsApp ekak karannam.", "මට ඔයාකේ number එක දාපලා කියලා, then WhatsApp එකක් කරන්නම්."),
    ("Neg_0005", "S", "Anekobo! suba welawak apage atulakda veevaa!", "අනෙකොබෝ! සුභ කවලාවක් අපකේ අටුලකද කව්වා!"),
    ("Neg_0006", "M", "Ayubowanda mata! kemathi innava? api apage wewedduwa thiyenne.", "ආයුබෝවන්ද මට! කෙමති ඉන්නවා? අපි අපකේ වවැඩදුව තිකයන්කන්."),
    ("Neg_0007", "S", "mata podak balanna puluwannam.", "මට පඩාක් බලන්න පුලුවන්නම්."),
    ("Neg_0008", "M", "oyata please eka mama gate enata msg ekak ewanna puluwan kada?", "ඔයාට please එක මම gate එනට msg එකක් එවන්න පුලුවන් කද?"),
    ("Neg_0009", "S", "ane, hari hari mama eka karala inna.", "අනේ, හරි හරි මම එක කරලා ඉන්න."),
    ("Neg_0010", "M", "Oyata ne, mama eka dunna ainee. Hari avurudu karana kiyala ainee.", "ඔයාට නේ, මම එක දුන්න අයිනේ. හරි අවුරුද කරන කියලා අයිනේ."),
    ("Neg_0011", "S", "Choti choti karanava, nathnam ammaru weyi.", "චෝටි චෝටි කරනවා, නැත්නම් අමාරු කවයි."),
    ("Neg_0012", "M", "Api api ekayi, mathak bala kota welata yamuda?", "අපි අපි එකයි, මතක බලා කොට කවලට යමුද?"),
    ("Neg_0013", "S", "mata eka denna... oyath balanna puluwan kada?", "මට එක දෙන්න... ඔයාත් බලන්න පුලුවන් කද?"),
    ("Neg_0014", "M", "ane! mama ayata katha karannam... kiyalaa oyata wadak hama?", "අනේ! මම අයාට තා කරන්නම්... කියලා ඔයාට වඩක් හම?"),
    ("Neg_0015", "S", "mang oyat passe katha karannam.", "මං ඔයාට පසේනසේ තා කරන්නම්."),
    ("Neg_0016", "M", "api ape rata adrae. oyage gama adrae.", "අපි අකේ රට ආදරෙයි. ඔයාකේ ගම ආදරෙයි."),
    ("Neg_0017", "S", "mata password eka forget karada?", "මට password එක forget කරාද?"),
    ("Neg_0018", "M", "oyage document eka email karannam, then signature liyanda add karanna.", "ඔයාකේ document එක email කරන්නම්, then signature ලියඳ add කරන්න."),
    ("Neg_0019", "S", "api as soon as possible yamuda?", "අපි as soon as possible යමුද?"),
    ("Neg_0020", "M", "mama office awe, kiyanna feeling tired very bad. oyata just wait karanna.", "මම office ආකව්, කියන්න feeling tired very bad. ඔයාට just wait කරන්න."),
    ("Neg_0021", "S", "oyage account eka logout karada?", "ඔයාකේ account එක logout කරාද?"),
    ("Neg_0022", "M", "mata download karanna puluwan kada? file format eka CSV walakura?", "මට download කරන්න පුලුවන් කද? file format එක CSV වලකුරා?"),
    ("Neg_0023", "S", "mata Instagram message ekak evanna.", "මට Instagram message එකක් එවන්න."),
    ("Neg_0024", "M", "Telegram eken notification thamai? Gmail check karada? mata notification ekat comment dennama.", "Telegram එන්කේ notification තමයි? Gmail check කරාද? මට notification එකට comment කදන්නම."),
    ("Neg_0025", "S", "mata COVID test eka karannam.", "මට COVID test එක කරන්නම්."),
    ("Neg_0026", "M", "oyage ID eka copy karada? PDF file eka convert karanna puluwanda?", "ඔයාකේ ID එක copy කරාද? PDF file එක convert කරන්න පුලුවන්ද?"),
    ("Neg_0027", "S", "mata uni eke exam ganna kalin study karanna.", "මට uni එකක් exam ගන්න කලින් study කරන්න."),
    ("Neg_0028", "M", "ape lab ekedi project ekak ganna kalin demo karanna karunakarala.", "අකේ lab එකක්දී project එකක් ගන්න කලින් demo කරන්න කරුණාකරලා."),
    ("Neg_0029", "S", "api Galle walata yamuda?", "අපි Galle වලට යමුද?"),
    ("Neg_0030", "M", "ada hawasata api Colombo Fort walata game karamu? Dematagoda eketin yamuda?", "අද වසට අපි Colombo Fort වලට game කරමු? Dematagoda එකතින් යමුද?"),
    ("Neg_0031", "S", "Sachini meeting ekedi veyda?", "සචිනි meeting එකක්දී කවයිද?"),
    ("Neg_0032", "M", "Pradeep saha Dilshan dennama project eka pooja kara? Jayasena eko pasal eka inna.", "ප්‍රදීප් සහ දිල්ශන් කදන්නම project එක පුජා කරා? ජයසෙන ඕකෝ පසල එක ඉන්න."),
    ("Neg_0033", "S", "mama exam ekedi 75n 100yama aragena.", "මම exam එකක්දී 75න් 100යම අරකගන."),
    ("Neg_0034", "M", "api 5 dina kola yamuda? 2025 awurudhe api apage planne karala.", "අපි 5 දිනක කෝල යමුද? 2025 අවුරුධෙ අපි අපකේ පලන්න කරලා."),
    ("Neg_0035", "S", "mata USD 100 denna puluwan kada?", "මට USD 100 දෙන්න පුලුවන් කද?"),
    ("Neg_0036", "M", "oyage salary eka Rs. 50000 kiyanne rupiyal lokuge best kada? GBP 2000 kiyala ayagena.", "ඔයාකේ salary එක Rs. 50000 කියන්කන් රුපියේ ලෝකුගේ best කද? GBP 2000 කියලා අයාගෙන."),
    ("Neg_0037", "S", "meeting eka 2:45PM ekata schedule karada?", "meeting එක 2:45PM එකට schedule කරාද?"),
    ("Neg_0038", "M", "api 08:30 AM ekata inna karunakarala. 6.00 PM walata class thamai.", "අපි 08:30 AM එකට ඉන්න කරුණාකරලා. 6.00 PM ෙලට class තමයි."),
    ("Neg_0039", "S", "appointment eka 2026-03-15 ekata booking karada?", "appointment එක 2026-03-15 එකට booking කරාද?"),
    ("Neg_0040", "M", "next holiday eka 5th May walata ayinaawa. birthday eka 14/02/2026 ekata celebrate karannam.", "next holiday එක 5th May ෙලට අයිනාවා. birthday එක 14/02/2026 එකට celebrate කරන්නම්."),
    ("Neg_0041", "S", "mama 2kg withara poth kanne.", "මම 2kg විතර පෝත් කන්කන්."),
    ("Neg_0042", "M", "e hadisi lama 5.5m withara lodata thiyenne. temperature 36.5celsius walakuta inna.", "එ හදිසි ලම 5.5m විතර ලෝදට තිකයන්කන්. temperature 36.5celsius වලකුට ඉන්න."),
    ("Neg_0043", "S", "sirawata machan, uba wedin karanava natha?", "සිරාවට චේ, උබ වැඩින් කරනවා නැතා?"),
    ("Neg_0044", "M", "bro eka kota halkari? mama eke set walakuta inna karunakarala. uba mara karana kotuwa!", "bro එක කොට හල්කරි? මම එකක් set වලකුට ඉන්න කරුණාකරලා. උබ මාර කරන කෝටුවා!"),
    ("Neg_0045", "S", "mata me link eka balanna: https://youtube.com/@channel", "මට නේ link එක බලන්න: https://youtube.com/@channel"),
    ("Neg_0046", "M", "website ekata @user123 kiyala message karannam. email: contact@domain.com evanna puluwanda?", "website එකට @user123 කියලා message කරන්නම්. email: contact@domain.com එවන්න පුලුවන්ද?"),
    ("Neg_0047", "S", "mama happy 😊 oyath kada?", "මම happy 😊 ඔයාත් කද?"),
    ("Neg_0048", "M", "uba very tired 😴 nathnam work karanna one. 🎯 target eka achieve karannam 💪", "උබ very tired 😴 නැත්නම් work කරන්න ඕනේ. 🎯 target එක achieve කරන්නම් 💪"),
    ("Neg_0049", "L", "api semester ekedi 4 subject thiyenne. Science, Maths, IT saha English. mama Science paper ekedi 85 arakagana. oyage subject thiyenna ada exam dates 2026-04-15, 2026-04-18, 2026-04-20, 2026-04-22 ayiya. Zoomekin study group ekak karamu. #ExamPrep #Study", "අපි semester එකක්දී 4 subject තිකයන්කන්. Science, Maths, IT සහ English. මම Science paper එකක්දී 85 අරකගන. ඔයාකේ subject තිකයන්න අද exam dates 2026-04-15, 2026-04-18, 2026-04-20, 2026-04-22 අයිශ. Zoomෙකින් study group එකක් කරමු. #ExamPrep #Study"),
    ("Neg_0050", "L", "machan mage graduation ceremony 2026-05-10 ekata thamai! #Proud #Family 🎓 oyata 7:30AM ekata awe kalin call karannam. mama blue suit ekat vesthaa, oyath kada? Havelock City reception ekakatai yamuda? oyage family samanya venue ekat pahala nisa transport arrange karannum. mata Rs.5000 withara budget thiyenne. kiyawada? 😅", "චේ මකේ graduation ceremony 2026-05-10 එකට තමයි! #Proud #Family 🎓 ඔයාට 7:30AM එකට ආකව් කලින් call කරන්නම්. මම blue suit එකට වෙස්තා, ඔයාත් කද? Havelock City reception එකට යමුද? ඔයාකේ family සමනය venue එකට පහලා නිසා transport arrange කරන්නුම. මට Rs.5000 විතර budget තිකයන්කන්. කියවද? 😅"),
]

def populate_excel(excel_path, test_cases_data):
    """Populate Excel file with test cases"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Unmerge all cells to avoid read-only issues
    for merged_cell_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell_range))
    
    # Clear existing data (keep headers in row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    
    # Populate test cases starting from row 2
    for idx, (tc_id, input_length, input_text, expected_output) in enumerate(test_cases_data, start=2):
        ws[f'A{idx}'] = tc_id
        ws[f'B{idx}'] = input_length
        ws[f'C{idx}'] = input_text
        ws[f'D{idx}'] = expected_output
        # Leave Actual output and Status empty for automation to fill
        ws[f'E{idx}'] = None
        ws[f'F{idx}'] = None
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 12
    
    # Set text wrapping for better readability
    for row in ws.iter_rows(min_row=2, max_row=len(test_cases_data)+1, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(excel_path)
    print(f"✓ Successfully populated {len(test_cases_data)} test cases into Excel file")
    print(f"✓ File saved: {excel_path}")

if __name__ == "__main__":
    excel_file = r"d:\test_automation\test_automation\Assignment 1 - Test cases.xlsx"
    populate_excel(excel_file, test_cases)
